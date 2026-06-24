#!/usr/bin/env python3
"""本地代理 - 提供AI群聊 + Excel数据同步 (端口8773)"""
import http.server, json, urllib.request, urllib.error, os, sys, msoffcrypto, tempfile
from openpyxl import load_workbook

PORT = 8773
EXCEL_PATH = r"D:\工作\烽火世游\密码管理.xlsx"
EXCEL_PWD = "lp26312lp"

def get_games_data():
    with open(EXCEL_PATH, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=EXCEL_PWD)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        office.decrypt(tmp)
        tmp.close()
        wb = load_workbook(tmp.name, data_only=True)
        ws = wb['游戏名称']
        records = []
        for r in range(3, ws.max_row + 1):
            uploader = ws.cell(r, 2).value
            game = ws.cell(r, 3).value
            pkg = ws.cell(r, 4).value
            backend = ws.cell(r, 5).value
            status = ws.cell(r, 6).value
            if not uploader or not game: continue
            skip = ['开发中','已下架','不跑了','已停量','没跑起来','没量了','已停投','已经下架']
            if status and str(status).strip() in skip: continue
            daily = []
            for c in range(7, 22):
                v = ws.cell(r, c).value
                if v is None or v == '' or v == 0: daily.append('-')
                elif isinstance(v, (float, int)): daily.append(str(round(float(v))))
                else: daily.append('-')
            records.append({
                'uploader': str(uploader).strip(),
                'game': str(game).strip(),
                'pkg': str(pkg).strip() if pkg else '',
                'backend': str(backend).strip() if backend else '',
                'status': str(status).strip() if status else '-',
                'data': daily
            })
        os.unlink(tmp.name)
        return records

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_cors()
        self.end_headers()

    def do_GET(self):
        if self.path == '/api/games-data':
            try:
                data = get_games_data()
                self.send_response(200)
                self.send_cors()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(data).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
        else:
            super().do_GET()

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length).decode('utf-8'))
        api_url = body.get('apiUrl', '').rstrip('/')
        if not api_url.endswith('/v1'): api_url += '/v1'
        target = api_url + '/chat/completions'
        payload = json.dumps({
            'model': body.get('model', ''),
            'messages': body.get('messages', []),
            'temperature': body.get('temperature', 0.7),
            'max_tokens': body.get('max_tokens', 2000)
        }).encode('utf-8')
        req = urllib.request.Request(target, data=payload,
            headers={'Content-Type': 'application/json', 'Authorization': 'Bearer ' + body.get('apiKey', '')})
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode('utf-8'))
            self.send_response(200); self.send_cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(result).encode('utf-8'))
        except Exception as e:
            self.send_response(500); self.send_cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            err = str(e.read().decode('utf-8', errors='replace')) if hasattr(e, 'read') else str(e)
            self.wfile.write(json.dumps({'error': err}).encode('utf-8'))

    def send_cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
    def log_message(self, *args): pass

if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    server = http.server.HTTPServer(('127.0.0.1', PORT), Handler)
    print(f'🚀 本地服务已启动: http://127.0.0.1:{PORT}')
    print(f'📊 游戏同步: http://127.0.0.1:{PORT}/games_sync.html')
    print(f'🤖 AI群聊:   http://127.0.0.1:{PORT}/ai_group.html')
    try: server.serve_forever()
    except KeyboardInterrupt: print('\n已停止')
